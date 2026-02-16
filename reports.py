import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Arial", bold=True, size=14)
SUBHEADER_FONT = Font(name="Arial", bold=True, size=12)
COLUMN_HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
CELL_FONT = Font(name="Arial", size=11)
AMOUNT_FONT = Font(name="Arial", bold=True, size=11)
TOTAL_FONT = Font(name="Arial", bold=True, size=13, color="1a5276")

HEADER_FILL = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="d5f5e3", end_color="d5f5e3", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def _style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = COLUMN_HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_data_row(ws, row, col_count, is_alt=False):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        if is_alt:
            cell.fill = ALT_ROW_FILL


def generate_expense_report(expenses: list, date_from: date, date_to: date, user_name: str = "") -> bytes:
    """Generate Excel expense report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Expense Report — {user_name}" if user_name else "Expense Report"
    ws["A1"].font = HEADER_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:E2")
    ws["A2"] = f"{date_from.strftime('%d %b %Y')} to {date_to.strftime('%d %b %Y')}"
    ws["A2"].font = SUBHEADER_FONT
    ws["A2"].alignment = CENTER

    # Column headers
    headers = ["#", "Date", "Category", "Description", "Amount (₹)"]
    col_widths = [6, 15, 20, 35, 18]
    header_row = 4

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=header_row, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    _style_header_row(ws, header_row, len(headers))

    # Data rows
    total = 0
    for idx, exp in enumerate(expenses):
        row = header_row + 1 + idx
        ws.cell(row=row, column=1, value=idx + 1).alignment = CENTER
        ws.cell(row=row, column=2, value=exp["date"])
        ws.cell(row=row, column=3, value=exp.get("category_name", ""))
        ws.cell(row=row, column=4, value=exp.get("description", ""))
        amt_cell = ws.cell(row=row, column=5, value=float(exp["amount"]))
        amt_cell.number_format = "#,##0.00"
        amt_cell.font = AMOUNT_FONT
        amt_cell.alignment = RIGHT
        _style_data_row(ws, row, len(headers), is_alt=(idx % 2 == 1))
        total += float(exp["amount"])

    # Total row
    total_row = header_row + 1 + len(expenses)
    ws.merge_cells(f"A{total_row}:D{total_row}")
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=1).alignment = RIGHT
    total_cell = ws.cell(row=total_row, column=5, value=total)
    total_cell.font = TOTAL_FONT
    total_cell.number_format = "#,##0.00"
    total_cell.alignment = RIGHT
    for col in range(1, 6):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = THIN_BORDER

    # Category summary sheet
    ws2 = wb.create_sheet("Category Summary")
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Category-wise Summary"
    ws2["A1"].font = HEADER_FONT
    ws2["A1"].alignment = CENTER

    cat_totals = {}
    for exp in expenses:
        cat = exp.get("category_name", "Uncategorized")
        cat_totals[cat] = cat_totals.get(cat, 0) + float(exp["amount"])

    cat_headers = ["Category", "Count", "Total (₹)"]
    for i, h in enumerate(cat_headers, 1):
        ws2.cell(row=3, column=i, value=h)
    _style_header_row(ws2, 3, 3)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 18

    cat_counts = {}
    for exp in expenses:
        cat = exp.get("category_name", "Uncategorized")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    for idx, (cat, amt) in enumerate(sorted(cat_totals.items(), key=lambda x: -x[1])):
        row = 4 + idx
        ws2.cell(row=row, column=1, value=cat)
        ws2.cell(row=row, column=2, value=cat_counts.get(cat, 0)).alignment = CENTER
        c = ws2.cell(row=row, column=3, value=amt)
        c.number_format = "#,##0.00"
        c.font = AMOUNT_FONT
        _style_data_row(ws2, row, 3, is_alt=(idx % 2 == 1))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_party_report(party_name: str, entries: list, date_from: date, date_to: date) -> bytes:
    """Generate Excel ledger report for a specific party/client."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Ledger — {party_name}"
    ws["A1"].font = HEADER_FONT
    ws["A1"].alignment = CENTER

    ws.merge_cells("A2:G2")
    ws["A2"] = f"{date_from.strftime('%d %b %Y')} to {date_to.strftime('%d %b %Y')}"
    ws["A2"].font = SUBHEADER_FONT
    ws["A2"].alignment = CENTER

    # Headers
    headers = ["#", "Date", "Type", "Item", "Qty", "Rate (₹)", "Amount (₹)"]
    col_widths = [6, 15, 18, 25, 10, 14, 18]
    header_row = 4

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.cell(row=header_row, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    _style_header_row(ws, header_row, len(headers))

    TYPE_LABELS = {
        "goods_sold": "Goods Sold",
        "payment_received": "Payment Received",
        "payment_made": "Payment Made",
        "goods_returned": "Goods Returned",
        "goods_taken": "Goods Taken",
    }

    payable = 0  # what party owes us
    receivable = 0  # what we owe party

    for idx, entry in enumerate(entries):
        row = header_row + 1 + idx
        ws.cell(row=row, column=1, value=idx + 1).alignment = CENTER
        ws.cell(row=row, column=2, value=entry["date"])
        ws.cell(row=row, column=3, value=TYPE_LABELS.get(entry["entry_type"], entry["entry_type"]))
        ws.cell(row=row, column=4, value=entry.get("item_name", ""))
        qty = entry.get("quantity")
        unit = entry.get("unit", "")
        ws.cell(row=row, column=5, value=f"{qty} {unit}".strip() if qty else "")
        rate = entry.get("rate")
        if rate:
            ws.cell(row=row, column=6, value=float(rate)).number_format = "#,##0.00"
        amt_cell = ws.cell(row=row, column=7, value=float(entry["amount"]))
        amt_cell.number_format = "#,##0.00"
        amt_cell.font = AMOUNT_FONT
        amt_cell.alignment = RIGHT
        _style_data_row(ws, row, len(headers), is_alt=(idx % 2 == 1))

        et = entry["entry_type"]
        amt = float(entry["amount"])
        if et in ("goods_sold",):
            payable += amt
        elif et in ("payment_received",):
            payable -= amt
        elif et in ("payment_made",):
            receivable += amt
        elif et in ("goods_returned",):
            payable -= amt
        elif et in ("goods_taken",):
            receivable -= amt

    # Summary
    summary_row = header_row + 2 + len(entries)
    ws.merge_cells(f"A{summary_row}:G{summary_row}")
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = HEADER_FONT

    labels = [
        ("Total Goods Sold / Payable by Party", payable),
        ("Total Payments Made / Receivable from Party", receivable),
        ("Net Balance (Party owes us)" if payable - receivable >= 0 else "Net Balance (We owe party)", abs(payable - receivable)),
    ]
    for i, (label, val) in enumerate(labels):
        row = summary_row + 1 + i
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row=row, column=1, value=label).font = SUBHEADER_FONT
        c = ws.cell(row=row, column=7, value=val)
        c.font = TOTAL_FONT
        c.number_format = "#,##0.00"
        c.alignment = RIGHT
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = TOTAL_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
